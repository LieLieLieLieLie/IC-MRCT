"""
utils.py
--------
Evaluation metrics, publication-quality visualization, and XLSX export.
"""

import logging
import os
import shutil
from typing import Dict, List, Sequence

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm
from matplotlib.gridspec import GridSpec, GridSpecFromSubplotSpec
from matplotlib.patches import Ellipse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

OURS_COLOR = "#FF6666"
OTHER_COLORS = ["#FFAA53", "#50CC55", "#00DDDD", "#3399FF", "#6666FF", "#9933FF"]
PALETTE = [OURS_COLOR] + OTHER_COLORS

C_SCRATCH = "#FFAA53"
C_PIT = "#50CC55"
C_AREA = "#3399FF"
C_TRANSIT = "#6666FF"
C_APPROACH = "#9933FF"
C_REPAIR = "#333333"

BLUE = "#007FFF"
RED = "#FF4F4F"
SEQ_CMAP = LinearSegmentedColormap.from_list("white_to_blue", ["#FFFFFF", BLUE])
DIV_CMAP = LinearSegmentedColormap.from_list("red_white_blue", [RED, "#FFFFFF", BLUE])

FONT_SIZE = 20
TICK_SIZE = 17
LEGEND_SIZE = 16
SUBTITLE_SIZE = 20

matplotlib.rcParams.update({
    "font.family": "serif",
    "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
    "mathtext.fontset": "stix",
    "axes.labelsize": FONT_SIZE,
    "axes.titlesize": SUBTITLE_SIZE,
    "xtick.labelsize": TICK_SIZE,
    "ytick.labelsize": TICK_SIZE,
    "legend.fontsize": LEGEND_SIZE,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "lines.linewidth": 2.5,
    "figure.dpi": 160,
    "savefig.dpi": 600,
    "savefig.bbox": "tight",
    "pdf.fonttype": 42,
    "ps.fonttype": 42,
})

METRIC_LABELS = {
    "n_defects": "Number of regions",
    "total_path_pts": "Path points",
    "repair_pts": "Repair points",
    "transit_pts": "Transit points",
    "total_dist_mm": "Total path length (mm)",
    "transit_dist_mm": "Transit length (mm)",
    "idle_ratio": "Idle ratio",
    "productive_ratio": "Productive ratio",
    "coverage_mean": "Mean coverage",
    "coverage_min": "Minimum coverage",
    "coverage_loss": "Coverage loss",
    "posture_change_mean_deg": "Mean posture change (deg)",
    "posture_change_max_deg": "Max posture change (deg)",
    "posture_smoothness": "Posture smoothness",
    "n_posture_jumps": "Posture jumps",
    "path_length_per_region": "Path length per region (mm)",
    "transit_per_region": "Transit length per region (mm)",
    "repair_density": "Repair density",
    "efficiency_score": "Composite efficiency",
    "priority_completion_cost": "Priority completion cost",
    "priority_tardiness": "Priority tardiness",
    "priority_satisfaction": "Priority satisfaction",
    "planning_objective": "Safety-first objective",
    "obstacle_violation": "Obstacle violation",
    "transit_obstacle_violation": "Transit obstacle violation",
    "clearance_penalty": "Clearance penalty",
    "transit_clearance_penalty": "Transit clearance penalty",
    "min_obstacle_clearance": "Min obstacle clearance (mm)",
    "min_transit_clearance": "Min transit clearance (mm)",
    "safe_transition_score": "Safe transition score",
    "robot_constraint_violation": "Robot constraint violation",
    "ik_violation": "IK reachability violation",
    "joint_limit_risk": "Joint-limit risk",
    "singularity_risk": "Singularity risk",
    "tool_collision_risk": "Tool collision risk",
    "body_collision_risk": "Body collision risk",
    "min_joint_margin": "Min joint-limit margin",
    "min_singularity_margin": "Min singularity margin",
}

LOWER_IS_BETTER = {
    "total_path_pts", "repair_pts", "transit_pts", "total_dist_mm",
    "transit_dist_mm", "idle_ratio", "coverage_loss",
    "posture_change_mean_deg", "posture_change_max_deg",
    "n_posture_jumps", "path_length_per_region", "transit_per_region",
    "priority_completion_cost", "priority_tardiness", "planning_objective",
    "obstacle_violation", "transit_obstacle_violation",
    "clearance_penalty", "transit_clearance_penalty",
    "robot_constraint_violation", "ik_violation", "joint_limit_risk",
    "singularity_risk", "tool_collision_risk", "body_collision_risk",
}


def compute_metrics(full_path: List[Dict], ordered_defects, local_trajs: List[List[Dict]], stats: Dict,
                    obstacles: List[Dict] = None, cfg=None) -> Dict:
    """Aggregate trajectory quality metrics."""
    cov_rates = []
    for defect, traj in zip(ordered_defects, local_trajs):
        repair = [(p["x"], p["y"]) for p in traj if p["zone"] == "repair"]
        if not repair:
            continue
        z = defect.zones
        xs = np.linspace(z.core_cx - z.core_ax, z.core_cx + z.core_ax, 25)
        ys = np.linspace(z.core_cy - z.core_ay, z.core_cy + z.core_ay, 25)
        samples = np.array([(x, y) for x in xs for y in ys if z.contains_core(float(x), float(y))], dtype=float)
        if len(samples) == 0:
            continue
        repair_xy = np.array(repair, dtype=float)
        fp = cfg.local_p.contact_footprint_width if cfg is not None else 6.0
        radius = fp * 0.5
        covered = sum(float(((repair_xy - sample) ** 2).sum(axis=1).min()) <= radius ** 2 for sample in samples)
        cov_rates.append(covered / len(samples))

    normals = np.array([[p["nx"], p["ny"], p["nz"]] for p in full_path])
    if len(normals) > 1:
        dots = np.clip((normals[:-1] * normals[1:]).sum(1), -1, 1)
        angles = np.degrees(np.arccos(dots))
        posture_mean = float(angles.mean())
        posture_max = float(angles.max())
    else:
        posture_mean = posture_max = 0.0

    n_regions = max(1, len(ordered_defects))
    total_dist = stats["total_dist_mm"]
    transit_dist = stats["transit_dist_mm"]
    idle_ratio = stats["idle_ratio"]
    coverage_mean = float(np.mean(cov_rates)) if cov_rates else 0.0
    coverage_min = float(np.min(cov_rates)) if cov_rates else 0.0
    productive_ratio = 1.0 - idle_ratio
    posture_smoothness = 1.0 / (1.0 + posture_mean + 0.25 * posture_max + stats["n_posture_jumps"])
    efficiency_score = coverage_mean * productive_ratio * posture_smoothness
    priority_completion_cost = _priority_completion_cost(ordered_defects)
    priority_tardiness = _priority_tardiness(ordered_defects)
    priority_satisfaction = 1.0 / (1.0 + priority_completion_cost)
    safety_margin = float(getattr(getattr(cfg, "transition", None), "obstacle_margin", 0.0)) if cfg is not None else 0.0
    obstacle_violation, min_clearance = _obstacle_metrics(
        full_path, obstacles or [], zones=("transit", "approach", "retract"), safety_margin=safety_margin
    )
    transit_violation, min_transit_clearance = _obstacle_metrics(
        full_path, obstacles or [], zones=("transit",), safety_margin=safety_margin
    )
    margin_scale = max(10.0, safety_margin)
    clearance_penalty = max(0.0, safety_margin - min_clearance) / margin_scale
    transit_clearance_penalty = max(0.0, safety_margin - min_transit_clearance) / margin_scale
    posture_penalty = min(2.5, posture_mean / 10.0 + posture_max / 45.0 + stats["n_posture_jumps"] / 8.0)
    safety_loss = 2.4 * transit_violation + 0.8 * obstacle_violation + 1.4 * transit_clearance_penalty + 0.35 * clearance_penalty
    safe_transition_score = 1.0 / (1.0 + safety_loss)
    robot_metrics = {}
    robot_violation = 0.0
    if cfg is not None:
        from robot_constraints import trajectory_robot_metrics
        robot_metrics = trajectory_robot_metrics(full_path, cfg, obstacles or [])
        robot_violation = robot_metrics.get("robot_constraint_violation", 0.0)
    planning_objective = (
        0.20 * idle_ratio
        + 0.12 * posture_penalty
        + 0.04 * stats["n_posture_jumps"]
        + 0.05 * (1.0 - coverage_min)
        + 0.01 * priority_completion_cost
        + 0.01 * priority_tardiness
        + 0.34 * transit_violation
        + 0.12 * obstacle_violation
        + 0.24 * transit_clearance_penalty
        + 0.06 * clearance_penalty
        + 0.10 * robot_violation
    )

    metrics = {
        "n_defects": len(ordered_defects),
        "total_path_pts": stats["total_points"],
        "repair_pts": stats["repair_points"],
        "transit_pts": stats["transit_points"],
        "total_dist_mm": total_dist,
        "transit_dist_mm": transit_dist,
        "idle_ratio": idle_ratio,
        "productive_ratio": productive_ratio,
        "coverage_mean": coverage_mean,
        "coverage_min": coverage_min,
        "coverage_loss": 1.0 - coverage_min,
        "posture_change_mean_deg": posture_mean,
        "posture_change_max_deg": posture_max,
        "posture_smoothness": posture_smoothness,
        "n_posture_jumps": stats["n_posture_jumps"],
        "path_length_per_region": total_dist / n_regions,
        "transit_per_region": transit_dist / n_regions,
        "repair_density": stats["repair_points"] / (total_dist + 1e-9),
        "efficiency_score": efficiency_score,
        "priority_completion_cost": priority_completion_cost,
        "priority_tardiness": priority_tardiness,
        "priority_satisfaction": priority_satisfaction,
        "obstacle_violation": obstacle_violation,
        "transit_obstacle_violation": transit_violation,
        "clearance_penalty": clearance_penalty,
        "transit_clearance_penalty": transit_clearance_penalty,
        "min_obstacle_clearance": min_clearance,
        "min_transit_clearance": min_transit_clearance,
        "safe_transition_score": safe_transition_score,
        "planning_objective": planning_objective,
    }
    metrics.update(robot_metrics)
    return metrics


def _priority_completion_cost(ordered_defects):
    if not ordered_defects:
        return 0.0
    priorities = np.array([int(d.priority) for d in ordered_defects], dtype=float)
    ranks = np.arange(1, len(priorities) + 1, dtype=float)
    actual = float(np.sum(priorities * ranks))
    ideal = float(np.sum(np.sort(priorities)[::-1] * ranks))
    worst = float(np.sum(np.sort(priorities) * ranks))
    if abs(worst - ideal) < 1e-12:
        return 0.0
    return (actual - ideal) / (worst - ideal)


def _priority_tardiness(ordered_defects):
    if not ordered_defects:
        return 0.0
    priorities = np.array([int(d.priority) for d in ordered_defects], dtype=float)
    ranks = np.arange(len(priorities), dtype=float)
    high = np.maximum(0.0, priorities - 1.0)
    if float(high.sum()) < 1e-12 or len(priorities) <= 1:
        return 0.0
    return float(np.sum(high * ranks) / ((len(priorities) - 1) * np.sum(high)))


def _obstacle_metrics(full_path, obstacles, zones=("transit", "approach", "retract"), safety_margin: float = 0.0):
    if not obstacles or len(full_path) < 2:
        return 0.0, 1e3
    segments = _zone_segments(full_path, zones)
    if not segments:
        segments = [np.array([[p["x"], p["y"]] for p in full_path], dtype=float)]
    violations = 0
    min_clearance = 1e3
    n_edges = 0
    for pts in segments:
        if len(pts) < 2:
            continue
        for p0, p1 in zip(pts[:-1], pts[1:]):
            n_edges += 1
            for obs in obstacles:
                clearance = _segment_circle_clearance(p0, p1, obs)
                min_clearance = min(min_clearance, clearance)
                if clearance < safety_margin:
                    violations += 1
    return violations / max(1, n_edges), float(min_clearance)


def _zone_segments(full_path, zones):
    trim_transit = tuple(zones) == ("transit",)
    segments = []
    i = 0
    while i < len(full_path):
        zone = full_path[i]["zone"]
        j = i + 1
        while j < len(full_path) and full_path[j]["zone"] == zone:
            j += 1
        if zone in zones:
            segment = full_path[i:j]
            if trim_transit and len(segment) > 4:
                segment = segment[2:-2]
            if len(segment) >= 2:
                segments.append(np.array([[p["x"], p["y"]] for p in segment], dtype=float))
        i = j
    return segments


def _points_for_zones(full_path, zones):
    trim_transit = tuple(zones) == ("transit",)
    pts = []
    i = 0
    while i < len(full_path):
        zone = full_path[i]["zone"]
        j = i + 1
        while j < len(full_path) and full_path[j]["zone"] == zone:
            j += 1
        if zone in zones:
            segment = full_path[i:j]
            if trim_transit and len(segment) > 4:
                segment = segment[2:-2]
            pts.extend([[p["x"], p["y"]] for p in segment])
        i = j
    return np.array(pts, dtype=float)


def _segment_circle_clearance(p0, p1, obs):
    c = np.array([obs["cx"], obs["cy"]], dtype=float)
    p0 = np.asarray(p0, dtype=float)
    p1 = np.asarray(p1, dtype=float)
    v = p1 - p0
    vv = float(np.dot(v, v))
    if vv < 1e-12:
        dist = float(np.linalg.norm(p0 - c))
    else:
        t = float(np.clip(np.dot(c - p0, v) / vv, 0.0, 1.0))
        dist = float(np.linalg.norm((p0 + t * v) - c))
    return dist - float(obs["r"])


def format_metrics(m: Dict, title: str = "") -> str:
    lines = ["-" * 62]
    if title:
        lines += [f"  {title}", "-" * 62]
    for k, v in m.items():
        val = f"{v:.4f}" if isinstance(v, float) else str(v)
        lines.append(f"  {k:<34} {val}")
    lines.append("-" * 62)
    return "\n".join(lines)


def save_experiment_xlsx(summary: Dict[str, Dict], records: List[Dict], save_path: str):
    """Save summary, standard deviation, and per-scene records to an Excel workbook."""
    os.makedirs(os.path.dirname(os.path.abspath(save_path)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    metrics = _metric_order(summary, records)
    _write_sheet(ws, ["Method"] + metrics, [[method] + [values.get(metric, "") for metric in metrics]
                                            for method, values in summary.items()])

    ws_std = wb.create_sheet("Std")
    _write_sheet(ws_std, ["Method"] + metrics, [[method] + [values.get(metric + "_std", "") for metric in metrics]
                                                for method, values in summary.items()])

    if records:
        ws_records = wb.create_sheet("PerScene")
        record_keys = ["scene_id", "method"] + metrics
        _write_sheet(ws_records, record_keys, [[row.get(key, "") for key in record_keys] for row in records])

    wb.save(save_path)
    logger.info("XLSX saved -> %s", save_path)


def save_summary_xlsx(summary: Dict[str, Dict], save_path: str):
    save_experiment_xlsx(summary, [], save_path)


def _write_sheet(ws, headers: Sequence[str], rows: Sequence[Sequence]):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(name="Times New Roman", bold=True, size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    ws.freeze_panes = "B2"
    for col in ws.columns:
        width = max(12, min(34, max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width


def _metric_order(summary: Dict[str, Dict], records: List[Dict]):
    preferred = [
        "n_defects", "total_path_pts", "repair_pts", "transit_pts",
        "total_dist_mm", "transit_dist_mm", "path_length_per_region",
        "transit_per_region", "idle_ratio", "productive_ratio",
        "coverage_mean", "coverage_min", "coverage_loss",
        "posture_change_mean_deg", "posture_change_max_deg",
        "posture_smoothness", "n_posture_jumps", "repair_density",
        "efficiency_score", "priority_completion_cost", "priority_tardiness",
        "priority_satisfaction", "obstacle_violation", "transit_obstacle_violation",
        "clearance_penalty", "transit_clearance_penalty",
        "min_obstacle_clearance", "min_transit_clearance", "safe_transition_score",
        "robot_constraint_violation", "ik_violation", "joint_limit_risk",
        "singularity_risk", "tool_collision_risk", "body_collision_risk",
        "min_joint_margin", "min_singularity_margin",
        "planning_objective",
    ]
    keys = set()
    for values in summary.values():
        keys.update(k for k in values if not k.endswith("_std"))
    for row in records:
        keys.update(k for k in row if k not in ("scene_id", "method"))
    return [k for k in preferred if k in keys] + sorted(k for k in keys if k not in preferred)


def _method_colors(methods):
    colors = []
    other_i = 0
    for method in methods:
        if method == "Ours":
            colors.append(OURS_COLOR)
        else:
            colors.append(OTHER_COLORS[other_i % len(OTHER_COLORS)])
            other_i += 1
    return colors


def _defect_color(defect):
    from defect import DefectType
    return {
        DefectType.SCRATCH: C_SCRATCH,
        DefectType.PIT: C_PIT,
        DefectType.AREA: C_AREA,
    }.get(defect.type, "#777777")


def _surface_grid(cfg, nx=170, ny=96):
    sc = cfg.surface
    xs = np.linspace(0, sc.length, nx)
    ys = np.linspace(0, sc.width, ny)
    X, Y = np.meshgrid(xs, ys)
    Z = sc.curvature_x * X ** 2 + sc.curvature_y * Y ** 2
    return X, Y, Z


def _heatmap_style(data):
    dmin = float(np.nanmin(data))
    dmax = float(np.nanmax(data))
    if dmin < 0 < dmax:
        vmax = max(abs(dmin), abs(dmax))
        return DIV_CMAP, TwoSlopeNorm(vmin=-vmax, vcenter=0.0, vmax=vmax)
    return SEQ_CMAP, None


def _bottom_legend(fig, handles, ncol):
    fig.legend(handles=handles, loc="lower center", ncol=ncol, frameon=False,
               bbox_to_anchor=(0.5, 0.012), columnspacing=1.6, handlelength=2.4)


def plot_scene_overview(defects, ordered_defects, cfg, save_path: str, obstacles: List[Dict] = None):
    fig = plt.figure(figsize=(16.5, 8.2), constrained_layout=False)
    gs = GridSpec(2, 2, height_ratios=[15, 1.35], hspace=0.22, wspace=0.16, figure=fig)
    ax0 = fig.add_subplot(gs[0, 0])
    ax1 = fig.add_subplot(gs[0, 1])
    cax0 = fig.add_subplot(gs[1, 0])
    cax1 = fig.add_subplot(gs[1, 1])
    X, Y, Z = _surface_grid(cfg)
    cmap, norm = _heatmap_style(Z)

    cf0 = ax0.contourf(X, Y, Z, levels=28, cmap=cmap, norm=norm)
    ax0.contour(X, Y, Z, levels=9, colors="#666666", linewidths=0.55, alpha=0.45)
    cb0 = fig.colorbar(cf0, cax=cax0, orientation="horizontal")
    cb0.ax.tick_params(labelsize=TICK_SIZE)
    cb0.set_label("Surface height (mm)", fontsize=FONT_SIZE)
    ax0.set_title("(a) Detected abnormal regions", pad=10)

    cf1 = ax1.contourf(X, Y, Z, levels=28, cmap=SEQ_CMAP, alpha=0.74)
    cb1 = fig.colorbar(cf1, cax=cax1, orientation="horizontal")
    cb1.ax.tick_params(labelsize=TICK_SIZE)
    cb1.set_label("Surface height (mm)", fontsize=FONT_SIZE)
    ax1.set_title("(b) Ours repair sequence", pad=10)

    for ax in (ax0, ax1):
        for obs in obstacles or []:
            ax.add_patch(plt.Circle((obs["cx"], obs["cy"]), obs["r"],
                                    facecolor="#222222", edgecolor="#000000",
                                    alpha=0.18, linewidth=1.2, hatch="///"))
        for defect in defects:
            z = defect.zones
            col = _defect_color(defect)
            ang = np.degrees(z.core_angle)
            ax.add_patch(Ellipse((z.core_cx, z.core_cy), 2 * z.safe_ax, 2 * z.safe_ay,
                                 angle=ang, facecolor=col, alpha=0.07,
                                 edgecolor=col, linewidth=1.2, linestyle=":"))
            ax.add_patch(Ellipse((z.core_cx, z.core_cy), 2 * z.trans_ax, 2 * z.trans_ay,
                                 angle=ang, facecolor=col, alpha=0.14,
                                 edgecolor=col, linewidth=1.4, linestyle="--"))
            ax.add_patch(Ellipse((z.core_cx, z.core_cy), 2 * z.core_ax, 2 * z.core_ay,
                                 angle=ang, facecolor=col, alpha=0.38,
                                 edgecolor=col, linewidth=1.9))

    centers = [(d.cx, d.cy) for d in ordered_defects]
    for k in range(len(centers) - 1):
        ax1.annotate("", xy=centers[k + 1], xytext=centers[k],
                     arrowprops=dict(arrowstyle="-|>", color=OURS_COLOR,
                                     lw=3.0, mutation_scale=21, shrinkA=8, shrinkB=8))
    for rank, defect in enumerate(ordered_defects, 1):
        ax1.text(defect.cx, defect.cy, str(rank), ha="center", va="center",
                 fontsize=16, fontweight="bold", color="white",
                 bbox=dict(boxstyle="circle,pad=0.30", facecolor=OURS_COLOR,
                           edgecolor="white", linewidth=1.0, alpha=0.95))

    handles = [
        mpatches.Patch(facecolor=C_SCRATCH, edgecolor=C_SCRATCH, alpha=0.75, label="Scratch"),
        mpatches.Patch(facecolor=C_PIT, edgecolor=C_PIT, alpha=0.75, label="Pit"),
        mpatches.Patch(facecolor=C_AREA, edgecolor=C_AREA, alpha=0.75, label="Area defect"),
        mpatches.Patch(facecolor="white", edgecolor="#555555", linestyle="--", label="Transition zone"),
        mpatches.Patch(facecolor="white", edgecolor="#555555", linestyle=":", label="Safety zone"),
        mpatches.Patch(facecolor="#222222", edgecolor="#000000", alpha=0.20, hatch="///", label="Forbidden zone"),
    ]
    _bottom_legend(fig, handles, ncol=6)

    for ax in (ax0, ax1):
        ax.set_xlim(0, cfg.surface.length)
        ax.set_ylim(0, cfg.surface.width)
        ax.set_xlabel("X (mm)")
        ax.set_ylabel("Y (mm)")
        ax.set_aspect("auto")
        ax.grid(True, color="#FFFFFF", linewidth=0.8, alpha=0.45)
    fig.subplots_adjust(bottom=0.16)
    _save(fig, save_path)


def plot_full_path(full_path: List[Dict], ordered_defects, cfg, save_path: str, obstacles: List[Dict] = None):
    FS  = FONT_SIZE + 2      # uniform enlarged font
    TFS = FS + 1             # subplot title: 1 pt above body

    fig = plt.figure(figsize=(16.5, 8.3), constrained_layout=False)
    gs = GridSpec(2, 2, height_ratios=[15, 1.35], hspace=0.34, wspace=0.07, figure=fig)
    ax0 = fig.add_subplot(gs[0, 0])
    ax1 = fig.add_subplot(gs[0, 1])
    cax0 = fig.add_subplot(gs[1, 0])
    cax1 = fig.add_subplot(gs[1, 1])
    X, Y, Z = _surface_grid(cfg)
    cf0 = ax0.contourf(X, Y, Z, levels=28, cmap=SEQ_CMAP, alpha=0.70)
    cb0 = fig.colorbar(cf0, cax=cax0, orientation="horizontal")
    cb0.ax.tick_params(labelsize=FS)
    cb0.set_label("Surface height (mm)", fontsize=FS)
    ax1.contourf(X, Y, Z, levels=28, cmap=SEQ_CMAP, alpha=0.55)

    zone_style = {
        "repair": dict(color=C_REPAIR, lw=1.9, ls="-", alpha=0.88),
        "transit": dict(color=C_TRANSIT, lw=2.5, ls="--", alpha=0.90),
        "approach": dict(color=C_APPROACH, lw=2.0, ls=":", alpha=0.92),
        "retract": dict(color=C_APPROACH, lw=2.0, ls=":", alpha=0.92),
    }
    i = 0
    while i < len(full_path) - 1:
        zone = full_path[i]["zone"]
        j = i + 1
        while j < len(full_path) and full_path[j]["zone"] == zone:
            j += 1
        seg_x = [p["x"] for p in full_path[i:j]]
        seg_y = [p["y"] for p in full_path[i:j]]
        if zone == "repair":
            mx, my = np.mean(seg_x), np.mean(seg_y)
            best_d = min(ordered_defects, key=lambda d: (d.cx - mx) ** 2 + (d.cy - my) ** 2)
            style = dict(zone_style[zone])
            style["color"] = _defect_color(best_d)
        else:
            style = zone_style.get(zone, zone_style["repair"])
        ax0.plot(seg_x, seg_y, **style)
        i = j

    pts = np.array([[p["x"], p["y"], p["z"]] for p in full_path])
    sc = ax1.scatter(pts[:, 0], pts[:, 1], c=pts[:, 2], s=8, cmap=SEQ_CMAP, edgecolors="none")
    cb1 = fig.colorbar(sc, cax=cax1, orientation="horizontal")
    cb1.ax.tick_params(labelsize=FS)
    cb1.set_label("Tool height (mm)", fontsize=FS)

    for ax in (ax0, ax1):
        for obs in obstacles or []:
            ax.add_patch(plt.Circle((obs["cx"], obs["cy"]), obs["r"],
                                    facecolor="#222222", edgecolor="#000000",
                                    alpha=0.16, linewidth=1.2, hatch="///"))
        for defect in ordered_defects:
            z = defect.zones
            ax.add_patch(Ellipse((z.core_cx, z.core_cy), 2 * z.core_ax, 2 * z.core_ay,
                                 angle=np.degrees(z.core_angle), fill=False,
                                 edgecolor=_defect_color(defect), linewidth=1.8))
        if full_path:
            ax.scatter([full_path[0]["x"]], [full_path[0]["y"]], s=145, marker="s",
                       color="#50CC55", edgecolor="white", linewidth=1.2, zorder=6)
            ax.scatter([full_path[-1]["x"]], [full_path[-1]["y"]], s=190, marker="*",
                       color=OURS_COLOR, edgecolor="white", linewidth=1.0, zorder=6)
        ax.set_xlim(0, cfg.surface.length)
        ax.set_ylim(0, cfg.surface.width)
        ax.set_xlabel("X (mm)", fontsize=FS)
        ax.tick_params(labelsize=FS)
        ax.set_aspect("auto")
        ax.grid(True, color="#FFFFFF", linewidth=0.8, alpha=0.40)
    ax0.set_ylabel("Y (mm)", fontsize=FS, labelpad=-8)
    ax1.set_ylabel("")
    ax1.tick_params(axis="y", left=False, labelleft=False)

    ax0.set_title("(a) Repair, approach, retract, and transit paths",
                  pad=14, fontsize=TFS, fontweight="bold")
    ax1.set_title("(b) Tool-height distribution along the trajectory",
                  pad=14, fontsize=TFS, fontweight="bold")

    handles = [
        mpatches.Patch(color=C_SCRATCH, label="Scratch repair"),
        mpatches.Patch(color=C_PIT, label="Pit repair"),
        mpatches.Patch(color=C_AREA, label="Area repair"),
        mpatches.Patch(color=C_TRANSIT, label="Transit"),
        mpatches.Patch(color=C_APPROACH, label="Approach/Retract"),
        mpatches.Patch(facecolor="#222222", edgecolor="#000000", alpha=0.20, hatch="///", label="Forbidden zone"),
        mpatches.Patch(color="#50CC55", label="Start"),
        mpatches.Patch(color=OURS_COLOR, label="End"),
    ]
    # 2-row centered legend (4 items per row)
    fig.legend(handles=handles, loc="lower center", ncol=4, frameon=False,
               fontsize=FS, bbox_to_anchor=(0.5, -0.02),
               columnspacing=1.6, handlelength=2.4)
    fig.subplots_adjust(bottom=0.22)
    _save(fig, save_path)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root, _ = os.path.splitext(save_path)
    for rel in [os.path.join("..", "paper", "IEEE", "figures"),
                os.path.join("..", "paper", "elsevier", "figures"),
                os.path.join("..", "paper", "figures")]:
        dest_dir = os.path.join(script_dir, rel)
        if os.path.isdir(dest_dir):
            shutil.copy2(root + ".pdf", os.path.join(dest_dir, os.path.basename(root + ".pdf")))


def plot_path_density_heatmap(full_path: List[Dict], cfg, save_path: str):
    repair_pts = np.array([[p["x"], p["y"]] for p in full_path if p["zone"] == "repair"])
    transit_pts = np.array([[p["x"], p["y"]] for p in full_path if p["zone"] == "transit"])
    all_pts = np.array([[p["x"], p["y"]] for p in full_path])
    if len(repair_pts) == 0:
        return
    heat, _, _ = np.histogram2d(
        repair_pts[:, 0], repair_pts[:, 1], bins=[130, 62],
        range=[[0, cfg.surface.length], [0, cfg.surface.width]],
    )
    positive = heat[heat > 0]
    if len(positive) > 0:
        heat = np.log1p(heat)
        vmax = float(np.percentile(heat[heat > 0], 98))
    else:
        vmax = 1.0

    fig = plt.figure(figsize=(15.2, 8.4), constrained_layout=False)
    gs = GridSpec(3, 1, height_ratios=[16, 1.10, 2.45], hspace=0.34, figure=fig)
    ax = fig.add_subplot(gs[0, 0])
    cax = fig.add_subplot(gs[1, 0])
    im = ax.imshow(heat.T, origin="lower", extent=[0, cfg.surface.length, 0, cfg.surface.width],
                   aspect="auto", cmap=SEQ_CMAP, interpolation="bilinear",
                   vmin=0, vmax=max(vmax, 1e-9))
    if len(all_pts) > 1:
        ax.plot(all_pts[:, 0], all_pts[:, 1], color="#333333", lw=0.75, alpha=0.32, label="Full trajectory")
    if len(transit_pts) > 1:
        ax.scatter(transit_pts[:, 0], transit_pts[:, 1], s=7, color=C_TRANSIT, alpha=0.55, label="Transit samples")
    x_centers = np.linspace(0, cfg.surface.length, heat.shape[0])
    y_centers = np.linspace(0, cfg.surface.width, heat.shape[1])
    ax.contour(x_centers, y_centers, heat.T, levels=5, colors=[BLUE], linewidths=0.7, alpha=0.55)
    cb = fig.colorbar(im, cax=cax, orientation="horizontal")
    cb.ax.tick_params(labelsize=TICK_SIZE)
    cb.set_label("Log-scaled repair path density", fontsize=FONT_SIZE)
    ax.set_title("(a) Repair path density with full trajectory overlay", pad=10)
    ax.set_xlabel("X (mm)")
    ax.set_ylabel("Y (mm)")
    handles = [
        mpatches.Patch(color="#333333", alpha=0.45, label="Full trajectory"),
        mpatches.Patch(color=C_TRANSIT, alpha=0.70, label="Transit samples"),
        mpatches.Patch(color=BLUE, alpha=0.70, label="High-density contours"),
    ]
    leg_ax = fig.add_subplot(gs[2, 0])
    leg_ax.axis("off")
    leg_ax.legend(handles=handles, loc="lower center", ncol=3, frameon=False,
                  bbox_to_anchor=(0.5, 0.08), columnspacing=2.2, handlelength=2.6)
    _save(fig, save_path)


def plot_posture_continuity(full_path: List[Dict], save_path: str):
    normals = np.array([[p["nx"], p["ny"], p["nz"]] for p in full_path])
    if len(normals) < 2:
        return
    dots = np.clip((normals[:-1] * normals[1:]).sum(1), -1, 1)
    angles = np.degrees(np.arccos(dots))
    idx    = np.arange(len(angles))

    zones  = [p.get("zone", "repair") for p in full_path[:-1]]

    window = max(5, len(angles) // 160)
    smooth = np.convolve(angles, np.ones(window) / window, mode="same")

    az = np.degrees(np.arctan2(normals[:, 0], normals[:, 2]))
    el = np.degrees(np.arcsin(np.clip(normals[:, 1], -1, 1)))

    ZONE_COLORS = {"repair": "#D0E8FF", "transit": "#FFE5CC", "approach": "#D6F5D6", "retract": "#D6F5D6"}
    ZONE_LABELS = {"repair": "Repair", "transit": "Transit", "approach": "App./Ret.", "retract": None}

    FS  = FONT_SIZE + 5
    TFS = FS + 6

    fig = plt.figure(figsize=(20, 13), constrained_layout=False)
    fig.subplots_adjust(left=0.07, right=0.97, top=0.93, bottom=0.08, hspace=0.52)
    # Top row spans full width; bottom row uses nested spec for tighter (b)–(c) gap
    gs_outer = fig.add_gridspec(2, 1, height_ratios=[1.35, 1.0],
                                 left=0.07, right=0.97, top=0.93, bottom=0.08,
                                 hspace=0.52)
    ax0 = fig.add_subplot(gs_outer[0])
    gs_bot = gs_outer[1].subgridspec(1, 2, wspace=0.22)
    ax2 = fig.add_subplot(gs_bot[0])
    ax3 = fig.add_subplot(gs_bot[1])

    # ── (a) time-series with segment shading ────────────────────────────────
    seen_labels = set()
    i = 0
    while i < len(zones):
        z = zones[i]
        j = i + 1
        while j < len(zones) and zones[j] == z:
            j += 1
        col   = ZONE_COLORS.get(z, "#EEEEEE")
        label = ZONE_LABELS.get(z, z)
        use_label = label if (label and label not in seen_labels) else None
        ax0.axvspan(i, j, color=col, alpha=0.55, lw=0, label=use_label)
        if use_label:
            seen_labels.add(use_label)
        i = j

    ax0.fill_between(idx, angles, color=OURS_COLOR, alpha=0.22)
    ax0.plot(idx, angles, color=OURS_COLOR, lw=1.2, alpha=0.75, label="Instantaneous $\Delta\psi$")
    ax0.plot(idx, smooth, color="#222222", lw=2.2, label="Moving average")

    ymax_data = angles.max() * 1.25
    y_ceil    = max(ymax_data, 0.5)
    ax0.set_ylim(0, y_ceil)
    ax0.axhline(15.0, color="#888888", lw=1.6, ls="--", label=r"$\theta_{\max}=15°$")
    pct95 = np.percentile(angles, 95)
    ax0.axhline(pct95, color="#CC4400", lw=1.4, ls=":", label=f"P95 = {pct95:.2f}°")

    # inset showing full scale
    from mpl_toolkits.axes_grid1.inset_locator import inset_axes
    axin = inset_axes(ax0, width="14%", height="68%", loc="upper right",
                      bbox_to_anchor=(-0.01, 0, 1, 1), bbox_transform=ax0.transAxes)
    axin.plot(idx, angles, color=OURS_COLOR, lw=0.8, alpha=0.6)
    axin.axhline(15.0, color="#888888", lw=1.2, ls="--")
    axin.set_ylim(0, 16)
    axin.set_xlim(0, len(angles))
    axin.set_ylabel("deg", fontsize=FS - 4)
    axin.tick_params(labelsize=FS - 5)
    axin.set_title("Full scale", fontsize=FS - 4, pad=3)
    axin.grid(True, axis="y", alpha=0.25)

    ax0.set_title("(a) Tool-normal change per step along the full trajectory",
                  pad=10, fontsize=TFS, fontweight="bold")
    ax0.set_xlabel("Path point index", fontsize=FS)
    ax0.set_ylabel("$\Delta\psi$ (deg)", fontsize=FS)
    ax0.tick_params(labelsize=FS - 1)
    handles, labels_leg = ax0.get_legend_handles_labels()
    ax0.legend(handles, labels_leg, loc="upper left", frameon=True,
               fontsize=FS - 2, ncol=3, framealpha=0.85)
    ax0.grid(True, axis="y", alpha=0.22)

    # ── (b) per-phase violin ─────────────────────────────────────────────────
    phase_map = {"repair": [], "transit": [], "approach": [], "retract": []}
    for ang, z in zip(angles, zones):
        phase_map.get(z, phase_map["repair"]).append(ang)
    app_ret = phase_map["approach"] + phase_map["retract"]
    groups  = [phase_map["repair"], phase_map["transit"], app_ret]
    glabels = ["Repair", "Transit", "App./Ret."]
    gcols   = [ZONE_COLORS["repair"], ZONE_COLORS["transit"], ZONE_COLORS["approach"]]
    gcols_e = ["#3399CC", "#CC6600", "#228822"]

    vp = ax2.violinplot(groups, positions=[1, 2, 3], showmedians=True,
                        showextrema=True, widths=0.6)
    for i_v, (body, ec) in enumerate(zip(vp["bodies"], gcols_e)):
        body.set_facecolor(gcols[i_v])
        body.set_edgecolor(ec)
        body.set_alpha(0.80)
    vp["cmedians"].set_color("#111111")
    vp["cmedians"].set_linewidth(2.2)
    for part in ("cbars", "cmins", "cmaxes"):
        vp[part].set_color("#555555")
        vp[part].set_linewidth(1.4)

    meds = [np.median(g) for g in groups if g]
    for xi, (med, g) in enumerate(zip(meds, groups), 1):
        ax2.text(xi, med + max(angles) * 0.04, f"{med:.3f}°",
                 ha="center", va="bottom", fontsize=FS - 3, color="#111111", fontweight="bold")

    ax2.set_xticks([1, 2, 3])
    ax2.set_xticklabels(glabels, fontsize=FS - 1)
    ax2.set_ylabel("$\Delta\psi$ (deg)", fontsize=FS)
    ax2.set_title("(b) Per-phase posture-change distribution",
                  pad=10, fontsize=TFS, fontweight="bold")
    ax2.tick_params(axis="y", labelsize=FS - 1)
    ax2.grid(True, axis="y", alpha=0.25)

    # ── (c) log-scale density histogram with percentile markers ─────────────
    max_val = angles.max()
    x_ceil  = max_val * 1.18
    bins = np.linspace(0, x_ceil, 55)
    counts, edges, patches = ax3.hist(angles, bins=bins, color=OURS_COLOR,
                                      alpha=0.82, edgecolor="white", linewidth=0.7,
                                      log=True)
    ax3.set_xlim(0, x_ceil)
    ax3.set_xlabel("$\Delta\psi$ (deg)", fontsize=FS)
    ax3.set_ylabel("Count (log scale)", fontsize=FS)
    ax3.set_title("(c) Posture-increment density (log scale)",
                  pad=10, fontsize=TFS, fontweight="bold")
    ax3.tick_params(labelsize=FS - 1)
    ax3.grid(True, axis="y", alpha=0.22)

    pct_vals = [(50, "#2266CC", "--"), (95, "#CC4400", ":"), (99, "#880000", "-.")]
    pct_labels = []
    for pct, col, ls in pct_vals:
        v = np.percentile(angles, pct)
        ax3.axvline(v, color=col, lw=1.8, ls=ls)
        pct_labels.append(mpatches.Patch(color=col, label=f"P{pct} = {v:.3f}°"))
    ax3.axvline(max_val, color="#550000", lw=1.8, ls="-")
    pct_labels.append(mpatches.Patch(color="#550000", label=f"Max = {max_val:.3f}°"))
    ax3.annotate(r"$\theta_{\max}=15°\!\gg\!$", xy=(x_ceil, 1),
                 xytext=(x_ceil * 0.72, 2.5), fontsize=FS - 4, color="#888888",
                 arrowprops=dict(arrowstyle="->", color="#888888", lw=1.2))
    ax3.legend(handles=pct_labels, fontsize=FS - 3, frameon=True, framealpha=0.85)

    # ── axis label for (b) becomes (c) if only 3 panels ────────────────────
    # Add a 4th annotation: re-label ax3 as (c), ax2 as (b) — already done above.

    _save(fig, save_path)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for rel in [os.path.join("..", "paper", "figures"),
                os.path.join("..", "paper", "IEEE", "figures"),
                os.path.join("..", "paper", "elsevier", "figures")]:
        dest = os.path.join(script_dir, rel)
        if os.path.isdir(dest):
            root, _ = os.path.splitext(save_path)
            shutil.copy2(root + ".pdf", os.path.join(dest, os.path.basename(root + ".pdf")))


def plot_comparison(results: Dict[str, Dict], save_path: str, metrics_to_plot: List[str] = None):
    if metrics_to_plot is None:
        metrics_to_plot = ["idle_ratio", "coverage_mean", "posture_change_mean_deg", "n_posture_jumps"]
    methods = list(results.keys())
    colors = _method_colors(methods)
    labels = [m.replace("Global-Local Planner", "Ours") for m in methods]
    n_met = len(metrics_to_plot)

    FS  = FONT_SIZE + 2      # uniform font size
    TFS = FS + 2             # subplot titles: 2 pt above body text

    n_methods = len(methods)
    # Wider subplots when many methods so bar annotations stay readable
    subplot_w  = 6.0 if n_methods <= 5 else 8.0
    bar_w      = 0.55 if n_methods <= 5 else 0.42
    ncol_leg   = n_methods          # always one row
    bot_margin = 0.13 if n_methods <= 5 else 0.15

    fig, axes = plt.subplots(1, n_met, figsize=(subplot_w * n_met, 6.2))
    axes = np.array(axes).reshape(-1)
    fig.subplots_adjust(left=0.07, right=0.98, bottom=bot_margin, top=0.88, wspace=0.24)

    x = np.arange(n_methods)
    for idx, (ax, metric) in enumerate(zip(axes, metrics_to_plot)):
        vals = [results[m].get(metric, 0) for m in methods]
        stds = [results[m].get(metric + "_std", 0) for m in methods]
        bars = ax.bar(x, vals, width=bar_w, color=colors, edgecolor="white", linewidth=1.2, zorder=3)
        ax.errorbar(x, vals, yerr=stds, fmt="none", ecolor="#333333",
                    capsize=4, elinewidth=1.5, zorder=4)
        ax.set_title(f"({chr(97 + idx)}) {METRIC_LABELS.get(metric, metric)}",
                     pad=14, fontsize=TFS, fontweight="bold")
        ax.set_xticks(x)
        ax.set_xticklabels([])
        ax.tick_params(axis="x", bottom=False)
        ax.set_ylabel(METRIC_LABELS.get(metric, metric), fontsize=FS)
        ax.tick_params(axis="y", labelsize=FS)
        ax.grid(True, axis="y", alpha=0.28, zorder=0)
        ymax = max(vals) if vals else 1.0
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.025 * (ymax + 1e-9),
                    f"{v:.3f}", ha="center", va="bottom", fontsize=FS - 2)

    # Shared bottom legend (2 rows when many methods)
    legend_handles = [mpatches.Patch(color=c, label=l) for c, l in zip(colors, labels)]
    fig.legend(handles=legend_handles, loc="lower center", ncol=ncol_leg,
               frameon=False, fontsize=FS,
               bbox_to_anchor=(0.5, 0.01),
               handlelength=1.6, columnspacing=1.2)

    # Save primary + mirror to IEEE/Elsevier paper figures/
    _save(fig, save_path)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root, _ = os.path.splitext(save_path)
    src = root + ".pdf"
    for rel in [os.path.join("..", "paper", "IEEE", "figures"),
                os.path.join("..", "paper", "elsevier", "figures"),
                os.path.join("..", "paper", "figures")]:
        dest_dir = os.path.join(script_dir, rel)
        if os.path.isdir(dest_dir):
            shutil.copy2(src, os.path.join(dest_dir, os.path.basename(src)))
            logger.info("Comparison figure mirrored -> %s", dest_dir)


def plot_performance_landscape(summary: Dict[str, Dict], records: List[Dict], save_path: str,
                               metrics: List[str] = None):
    if metrics is None:
        metrics = [
            "idle_ratio", "transit_per_region", "path_length_per_region",
            "coverage_mean", "posture_change_mean_deg", "posture_change_max_deg",
            "posture_smoothness", "priority_satisfaction", "safe_transition_score",
            "transit_clearance_penalty", "planning_objective",
        ]
    methods = list(summary.keys())
    colors = _method_colors(methods)
    labels = [m.replace("Global-Local Planner", "Ours") for m in methods]
    score = _normalized_score_matrix(summary, methods, metrics)  # (n_methods, n_metrics)
    score_T = score.T                   # (n_metrics, n_methods) – rows=metrics, cols=methods

    FS   = FONT_SIZE          # single uniform font size for all text
    TFS  = FS + 1             # subplot title (1 pt larger)
    CELL = max(FS - 4, 10)   # cell annotation (smaller to fit in cells)

    def _compact(v: float) -> str:
        s = f"{v:.3f}".rstrip("0").rstrip(".")
        return s if s else "0"

    # ── Layout: heatmap (row 0, full width) + 3 subplots (row 1) ─────────────
    fig = plt.figure(figsize=(20, 11))
    gs = GridSpec(2, 3, figure=fig,
                  height_ratios=[1.0, 1.25],
                  left=0.10, right=0.98,
                  top=0.94, bottom=0.12,
                  hspace=0.58, wspace=0.28)
    ax_heat    = fig.add_subplot(gs[0, :])
    ax_line    = fig.add_subplot(gs[1, 0])
    ax_box     = fig.add_subplot(gs[1, 1])
    ax_scatter = fig.add_subplot(gs[1, 2])

    # ── (a) Heatmap – methods on x-axis, metrics on y-axis ───────────────────
    im = ax_heat.imshow(score_T, cmap=SEQ_CMAP, vmin=0, vmax=1, aspect="auto")
    ax_heat.set_title("(a) Normalized multi-metric performance map", pad=8, fontsize=TFS, fontweight="bold")
    # x-axis: method names in black (uniform color)
    ax_heat.set_xticks(np.arange(len(methods)))
    ax_heat.set_xticklabels(labels, fontsize=FS, color="black")
    # y-axis: metric short names, horizontal
    ax_heat.set_yticks(np.arange(len(metrics)))
    ax_heat.set_yticklabels([_short_label(m) for m in metrics],
                             rotation=0, ha="right", fontsize=FS)
    # Cell annotations
    for i in range(score_T.shape[0]):
        for j in range(score_T.shape[1]):
            v = score_T[i, j]
            ax_heat.text(j, i, _compact(v), ha="center", va="center",
                         fontsize=CELL,
                         color="#111111" if v < 0.65 else "white")
    # Colorbar via inset_axes – aligned exactly with heatmap left/right edges
    cax = ax_heat.inset_axes([0, -0.23, 1.0, 0.06])
    cb = fig.colorbar(im, cax=cax, orientation="horizontal")
    cb.ax.tick_params(labelsize=FS)
    cb.set_label("Normalized score (higher is better)", fontsize=FS)

    # ── (b) Line chart ────────────────────────────────────────────────────────
    radar_metrics = ["coverage_mean", "productive_ratio", "posture_smoothness",
                     "priority_satisfaction", "safe_transition_score", "efficiency_score"]
    radar = _normalized_score_matrix(summary, methods, radar_metrics)
    x_r = np.arange(len(radar_metrics))
    draw_order = ([i for i, m in enumerate(methods) if m != "Ours"] +
                  [i for i, m in enumerate(methods) if m == "Ours"])
    for i in draw_order:
        is_ours = methods[i] == "Ours"
        ax_line.plot(x_r, radar[i], color=colors[i],
                     lw=3.2 if is_ours else 1.8,
                     marker="o", markersize=7 if is_ours else 4.5,
                     zorder=5 if is_ours else 3)
    ax_line.set_title("(b) Planning quality comparison", pad=8, fontsize=TFS, fontweight="bold")
    ax_line.set_xticks(x_r)
    ax_line.set_xticklabels([_short_label(m) for m in radar_metrics],
                             rotation=30, ha="right", fontsize=FS)
    ax_line.set_ylim(0, 1.08)
    ax_line.set_ylabel("Normalized score", fontsize=FS)
    ax_line.yaxis.set_major_locator(plt.MultipleLocator(0.25))
    ax_line.tick_params(axis="y", labelsize=FS)
    ax_line.grid(True, alpha=0.28)

    # ── (c) Violin – no x-axis labels ────────────────────────────────────────
    box_metric = "idle_ratio"
    data = [[r[box_metric] for r in records if r["method"] == m] for m in methods]
    parts = ax_box.violinplot(data, showmeans=False, showextrema=False)
    for body, col in zip(parts["bodies"], colors):
        body.set_facecolor(col)
        body.set_edgecolor("white")
        body.set_alpha(0.62)
    bp = ax_box.boxplot(data, widths=0.18, patch_artist=True, showfliers=False)
    for patch, col in zip(bp["boxes"], colors):
        patch.set_facecolor(col)
        patch.set_alpha(0.88)
        patch.set_edgecolor("#333333")
    ax_box.set_title("(c) Idle ratio distribution", pad=8, fontsize=TFS, fontweight="bold")
    ax_box.set_xticks(np.arange(1, len(methods) + 1))
    ax_box.set_xticklabels([])
    ax_box.tick_params(axis="x", bottom=False)
    ax_box.set_ylabel("Idle ratio", fontsize=FS)
    ax_box.tick_params(axis="y", labelsize=FS)
    ax_box.grid(True, axis="y", alpha=0.25)

    # ── (d) Scatter ──────────────────────────────────────────────────────────
    for method, col, label in zip(methods, colors, labels):
        xs = [r["idle_ratio"] for r in records if r["method"] == method]
        ys = [r["posture_change_max_deg"] for r in records if r["method"] == method]
        ss = [30 + 160 * max(0, r["coverage_mean"]) for r in records if r["method"] == method]
        ax_scatter.scatter(xs, ys, s=ss, color=col, alpha=0.62,
                           edgecolor="white", linewidth=0.7)
    ax_scatter.set_title("(d) Efficiency–smoothness trade-off", pad=8, fontsize=TFS, fontweight="bold")
    ax_scatter.set_xlabel("Idle ratio", fontsize=FS)
    ax_scatter.set_ylabel("Max posture change (deg)", fontsize=FS)
    ax_scatter.tick_params(labelsize=FS)
    ax_scatter.grid(True, alpha=0.25)

    # ── Shared bottom legend ──────────────────────────────────────────────────
    legend_handles = [mpatches.Patch(color=c, label=l) for c, l in zip(colors, labels)]
    ncol = min(len(methods), 4)
    fig.legend(handles=legend_handles, loc="lower center", ncol=ncol,
               frameon=False, fontsize=FS,
               bbox_to_anchor=(0.5, -0.07),
               handlelength=2.0, columnspacing=1.8)

    # Save primary path then mirror to IEEE and Elsevier paper figures/
    _save(fig, save_path)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root, _ = os.path.splitext(save_path)
    src = root + ".pdf"
    for rel in [os.path.join("..", "paper", "IEEE", "figures"),
                os.path.join("..", "paper", "elsevier", "figures"),
                os.path.join("..", "paper", "figures")]:
        dest_dir = os.path.join(script_dir, rel)
        if os.path.isdir(dest_dir):
            shutil.copy2(src, os.path.join(dest_dir, os.path.basename(src)))
            logger.info("Landscape figure mirrored -> %s", dest_dir)


def _normalized_score_matrix(summary, methods, metrics):
    raw = np.array([[summary[m].get(metric, 0.0) for metric in metrics] for m in methods], dtype=float)
    score = np.zeros_like(raw)
    for j, metric in enumerate(metrics):
        col = raw[:, j]
        lo, hi = float(np.min(col)), float(np.max(col))
        if abs(hi - lo) < 1e-12:
            score[:, j] = 1.0
        elif metric in LOWER_IS_BETTER:
            score[:, j] = (hi - col) / (hi - lo)
        else:
            score[:, j] = (col - lo) / (hi - lo)
    return score


def _short_label(metric):
    labels = {
        "idle_ratio": "Idle",
        "transit_per_region": "Transit/reg.",
        "path_length_per_region": "Length/reg.",
        "coverage_mean": "Coverage",
        "productive_ratio": "Productive",
        "posture_change_mean_deg": "Mean posture",
        "posture_change_max_deg": "Max posture",
        "posture_smoothness": "Smoothness",
        "efficiency_score": "Efficiency",
        "repair_density": "Density",
        "priority_satisfaction": "Priority",
        "priority_tardiness": "Tardiness",
        "planning_objective": "Objective",
        "safe_transition_score": "Safety",
        "obstacle_violation": "Obstacle",
        "transit_obstacle_violation": "Transit obs.",
        "clearance_penalty": "Clearance",
        "transit_clearance_penalty": "Transit clear.",
        "robot_constraint_violation": "Robot constr.",
        "ik_violation": "IK",
        "joint_limit_risk": "Joint limit",
        "singularity_risk": "Singularity",
        "tool_collision_risk": "Tool coll.",
        "body_collision_risk": "Body coll.",
        }
    return labels.get(metric, METRIC_LABELS.get(metric, metric))


def _save(fig, path: str):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    root, _ = os.path.splitext(path)
    pdf_path = root + ".pdf"
    fig.savefig(pdf_path, format="pdf", bbox_inches="tight")
    plt.close(fig)
    logger.info("PDF figure saved -> %s", pdf_path)
